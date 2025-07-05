import tkinter as tk
from tkinter import ttk, messagebox, filedialog, font
import pythoncom
import win32com.client
from win32com.client import constants
import win32gui
import win32con
import time
import openpyxl
import os
from datetime import datetime
import threading
from PIL import Image, ImageTk
from excel_session_manager_link_updater import run_excel_link_update
import psutil
import gc

def kill_zombie_excel_processes(whitelist_hwnds=None):
    if whitelist_hwnds is None:
        whitelist_hwnds = set()
    zombie_pids = []
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'] and proc.info['name'].lower() == 'excel.exe':
            try:
                hwnds = []
                def callback(hwnd, hwnds):
                    _, pid = win32process.GetWindowThreadProcessId(hwnd)
                    if pid == proc.info['pid']:
                        hwnds.append(hwnd)
                win32gui.EnumWindows(callback, hwnds)
                keep = False
                for hwnd in hwnds:
                    if hwnd in whitelist_hwnds:
                        keep = True
                        break
                if not keep and not hwnds:
                    zombie_pids.append(proc.info['pid'])
            except Exception:
                continue
    for pid in zombie_pids:
        try:
            p = psutil.Process(pid)
            p.kill()
        except Exception:
            continue

MONO_FONTS = [
    "Consolas", "Courier New", "Fira Code", "JetBrains Mono",
    "DejaVu Sans Mono", "Source Code Pro", "Monaco"
]

def calc_row_height(fsize):
    return int(fsize * 2.1)

def calc_col2_width(fsize):
    return int(fsize * 11.5)

class DragSelectTreeview(ttk.Treeview):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._drag_start = None
        self._drag_mode = None
        self.bind("<Button-1>", self._on_click)
        self.bind("<B1-Motion>", self._on_drag)
        self.bind("<ButtonRelease-1>", self._on_release)

    def _on_click(self, event):
        iid = self.identify_row(event.y)
        if not iid:
            self.selection_remove(self.selection())
            return
        self.focus(iid)
        if iid in self.selection():
            self.selection_remove(iid)
            self._drag_mode = "unselect"
        else:
            self.selection_add(iid)
            self._drag_mode = "select"
        self._drag_start = iid
        return "break"

    def _on_drag(self, event):
        iid = self.identify_row(event.y)
        if not iid or not self._drag_start:
            return
        iids = self.get_children()
        start_idx = iids.index(self._drag_start)
        cur_idx = iids.index(iid)
        lo = min(start_idx, cur_idx)
        hi = max(start_idx, cur_idx)
        if self._drag_mode == "unselect":
            for i in iids[lo:hi+1]:
                self.selection_remove(i)
        else:
            for i in iids[lo:hi+1]:
                self.selection_add(i)
        return "break"

    def _on_release(self, event):
        self._drag_start = None
        self._drag_mode = None
        self.event_generate("<<TreeviewSelect>>")
        return "break"

def get_file_mtime_str(path):
    if os.path.exists(path):
        ts = os.path.getmtime(path)
        dt = datetime.fromtimestamp(ts)
        return f"{dt.year}-{dt.month:02d}-{dt.day:02d} {dt.hour:02d}:{dt.minute:02d}:{dt.second:02d}"
    else:
        return ''

def parse_mtime(mtime_str):
    try:
        parts = mtime_str.split()
        date, time_part = parts[0], parts[1]
        day, month, year = map(int, date.split('/'))
        hour, minute = map(int, time_part.split(':'))
        return datetime(year, month, day, hour, minute)
    except Exception:
        return None

class ConsolePopup(tk.Toplevel):
    def __init__(self, parent, title="Console Output"):
        super().__init__(parent)
        self.title(title)
        self.geometry("1400x1200")
        self.resizable(True, True)
        frm = tk.Frame(self)
        frm.pack(fill="both", expand=True, padx=8, pady=8)
        self.text = tk.Text(frm, font=("Consolas", 11), state="disabled", wrap="word", bg="#111", fg="#f9f9f9")
        self.text.pack(side="left", fill="both", expand=True)
        yscrollbar = tk.Scrollbar(frm, command=self.text.yview)
        yscrollbar.pack(side="right", fill="y")
        self.text["yscrollcommand"] = yscrollbar.set
        self.text.config(undo=True, autoseparators=True, maxundo=-1)
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.closed = False

    def print(self, msg):
        if self.closed:
            return
        import datetime
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.text.config(state="normal")
        self.text.insert("end", f"{ts} | {msg}\n")
        self.text.see("end")
        self.text.config(state="disabled")
        self.update_idletasks()

    def on_close(self):
        self.closed = True
        self.destroy()

class ExcelSessionManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Session Manager")
        self.root.geometry("1200x700")
        self.last_log_dir = r"D:\Pzone\Log"
        self.root.minsize(600, 500)
        self.root.protocol("WM_DELETE_WINDOW", lambda: self.root.destroy())
        self.is_mini = False
        self.normal_geometry = "1200x620"
        self.mini_side = 180
        self.mini_geometry = f"{self.mini_side}x{self.mini_side}+150+150"
        self.floating_icon_btn = None
        self.setup_ui()
        self.show_names()

    def setup_ui(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill="both")

        frame_session = ttk.Frame(self.notebook)
        self.notebook.add(frame_session, text="Session Manager")

        font_frame = tk.Frame(frame_session)
        font_frame.pack(pady=(5, 0), padx=10, anchor="nw", fill='x')

        self.pin_btn = tk.Button(
            font_frame,
            text="Mini Widget (Floating Pin)",
            font=("Microsoft JhengHei", 10),
            command=self.enter_mini,
            relief="raised",
            borderwidth=1,
            padx=5,
            pady=2
        )
        self.pin_btn.pack(side=tk.RIGHT, padx=(10, 0))

        tk.Label(font_frame, text="Font:").pack(side=tk.LEFT)
        self.font_family_var = tk.StringVar(value=MONO_FONTS[0])
        self.font_size_var = tk.IntVar(value=12)
        font_family_menu = ttk.OptionMenu(font_frame, self.font_family_var, MONO_FONTS[0], *MONO_FONTS, command=self.update_treeview_font)
        font_family_menu.pack(side=tk.LEFT, padx=(5, 15))

        tk.Label(font_frame, text="Size:").pack(side=tk.LEFT)
        font_size_spin = ttk.Spinbox(font_frame, from_=8, to=32, textvariable=self.font_size_var, width=4, command=self.update_treeview_font)
        font_size_spin.pack(side=tk.LEFT)
        self.font_size_var.trace_add("write", lambda *a: self.update_treeview_font())

        title_frame = tk.Frame(frame_session)
        title_frame.pack(pady=(10, 0), padx=10, anchor="w", fill='x')

        main_label = tk.Label(title_frame, text="Active Excel Session:", font=("Arial", 16, "bold"))
        main_label.pack(side=tk.LEFT)

        self.count_label = tk.Label(title_frame, text="", font=("Arial", 12))
        self.count_label.pack(side=tk.LEFT, padx=(5, 0), pady=(2,0))

        main_frame = tk.Frame(frame_session)
        main_frame.pack(pady=(5, 10), padx=10, expand=True, fill='both')

        listbox_frame = tk.Frame(main_frame)
        listbox_frame.pack(side=tk.LEFT, fill='both', expand=True, pady=10)

        select_console_frame = tk.Frame(listbox_frame)
        select_console_frame.pack(anchor='w', padx=0, pady=(0, 2), fill='x')

        self.select_all_var = tk.BooleanVar(value=False)
        select_all_cb = tk.Checkbutton(
            select_console_frame,
            text="Select All",
            variable=self.select_all_var,
            command=self.on_select_all_toggle
        )
        select_all_cb.pack(side='left', padx=(8,0))

        self.show_console_progress_var = tk.BooleanVar(value=True)
        show_console_cb = tk.Checkbutton(
            select_console_frame,
            text="Show Progress Console",
            variable=self.show_console_progress_var
        )
        show_console_cb.pack(side='left', padx=(12,0))

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Consolas", 12, "bold"))
        style.configure("Treeview", font=("Consolas", 12), rowheight=calc_row_height(self.font_size_var.get()))

        self.tree = DragSelectTreeview(listbox_frame, columns=("col1", "col2"), show="headings", selectmode="extended")
        self.tree.column("col1", anchor="w")
        self.tree.column("col2", anchor="e", width=320, stretch=False)
        tree_yscrollbar = tk.Scrollbar(listbox_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_yscrollbar.set)
        tree_yscrollbar.pack(side="right", fill="y")
        self.tree.pack(expand=True, fill='both', padx=8, pady=8)
        self.root.after(200, lambda: self.tree.xview_moveto(1.0))
        self.tree.bind("<MouseWheel>", lambda event: self.tree.yview_scroll(-1 * int(event.delta / 120), "units"))
        self.tree.bind("<Button-4>", lambda event: self.tree.yview_scroll(-1, "units"))
        self.tree.bind("<Button-5>", lambda event: self.tree.yview_scroll(1, "units"))
        self.tree.bind('<<TreeviewSelect>>', self.on_selection_change)
        self.tree.bind("<Button-1>", self.on_treeview_heading_click, add="+")

        side_btn_frame = tk.Frame(main_frame, width=120)
        side_btn_frame.pack(side=tk.RIGHT, padx=(10, 0), pady=(40, 0), fill='y')
        side_btn_frame.pack_propagate(False)
        btn_props = {'width': 20, 'height': 2, 'wraplength': 140, 'font': ("Arial", 10, "bold")}

        self.refresh_btn = tk.Button(side_btn_frame, text="Refresh\nList", **btn_props, command=self.show_names)
        self.refresh_btn.pack(pady=5, anchor='n')

        self.show_path_btn = tk.Button(side_btn_frame, text="Show\nFile Path", **btn_props, command=self.toggle_path)
        self.show_path_btn.pack(pady=5, anchor='n')

        activate_btn = tk.Button(side_btn_frame, text="Activate\nSelected", **btn_props, command=self.activate_selected_workbooks)
        activate_btn.pack(pady=5, anchor='n')

        minimize_btn = tk.Button(side_btn_frame, text="Minimize All\nExcel", **btn_props, command=self.minimize_all_excel)
        minimize_btn.pack(pady=5, anchor='n')

        save_sess_btn = tk.Button(side_btn_frame, text="Save\nSession", **btn_props, command=self.save_session)
        save_sess_btn.pack(pady=5, anchor='n')

        load_sess_btn = tk.Button(side_btn_frame, text="Load\nSession", **btn_props, command=self.load_session)
        load_sess_btn.pack(pady=5, anchor='n')

        update_links_btn = tk.Button(side_btn_frame, text="Update Recent\nExternal Links", **btn_props, command=self.open_link_update_options)
        update_links_btn.pack(pady=5, anchor='n')
        
        btn1 = tk.Button(side_btn_frame, text="Save\nSelected", **btn_props, command=self.save_selected_workbooks)
        btn1.pack(pady=5, anchor='n')

        btn2 = tk.Button(side_btn_frame, text="Close\nWithout Saving", **btn_props, command=lambda: self.close_selected_workbooks(False))
        btn2.pack(pady=5, anchor='n')

        btn3 = tk.Button(side_btn_frame, text="Save and Close\nSelected", **btn_props, command=lambda: self.close_selected_workbooks(True))
        btn3.pack(pady=5, anchor='n')

        self.file_names, self.file_paths, self.sheet_names, self.active_cells = [], [], [], []
        self.showing_path = False
        self.target_captions = []
        self.sort_states = {"col1": "none", "col2": "none"}
        self.original_data = []

    def on_select_all_toggle(self):
        if self.select_all_var.get():
            for iid in self.tree.get_children():
                self.tree.selection_add(iid)
        else:
            self.tree.selection_remove(self.tree.selection())

    def enter_mini(self):
        if self.is_mini:
            return
        self.is_mini = True
        self.normal_geometry = self.root.geometry()
        self.notebook.pack_forget()
        self.root.minsize(self.mini_side, self.mini_side)
        self.root.geometry(self.mini_geometry)
        self.root.resizable(False, False)
        self.root.wm_attributes("-topmost", 1)

        icon_path = "maximize_full_screen.png"
        icon_size = (self.mini_side - 40, self.mini_side - 40)
        if os.path.exists(icon_path):
            try:
                pil_image = Image.open(icon_path).resize(icon_size, Image.LANCZOS)
                icon_image = ImageTk.PhotoImage(pil_image)
                self.floating_icon_btn = tk.Button(
                    self.root, image=icon_image, width=icon_size[0], height=icon_size[1],
                    command=self.exit_mini, borderwidth=0, relief="flat"
                )
                self.floating_icon_btn.image = icon_image
            except Exception:
                self.create_fallback_icon()
        else:
            self.create_fallback_icon()
        self.floating_icon_btn.pack(expand=True, padx=10, pady=10)

    def exit_mini(self):
        if not self.is_mini:
            return
        self.is_mini = False
        if self.floating_icon_btn:
            self.floating_icon_btn.destroy()
            self.floating_icon_btn = None
        self.root.minsize(600, 500)
        self.root.geometry(self.normal_geometry)
        self.root.resizable(True, True)
        self.root.wm_attributes("-topmost", 0)
        self.notebook.pack(expand=True, fill="both")

    def create_fallback_icon(self):
        self.floating_icon_btn = tk.Button(
            self.root, text="🗔", font=("Segoe UI Emoji", 48),
            command=self.exit_mini, borderwidth=0, relief="flat"
        )

    def update_treeview_font(self, *args):
        new_font = (self.font_family_var.get(), self.font_size_var.get())
        style = ttk.Style()
        style.configure("Treeview.Heading", font=(self.font_family_var.get(), self.font_size_var.get(), "bold"))
        self.tree.tag_configure('custom_font', font=new_font)
        for iid in self.tree.get_children():
            self.tree.item(iid, tags=('custom_font',))
        row_h = calc_row_height(self.font_size_var.get())
        style.configure("Treeview", rowheight=row_h)
        test_str = "2025-06-29 11:32:48"
        test_font = font.Font(family=self.font_family_var.get(), size=self.font_size_var.get())
        col2_width = test_font.measure(test_str) + 18
        self.tree.column("col2", width=col2_width, stretch=False)

    def get_open_excel_files(self):
        pythoncom.CoInitialize()
        excel_files, file_paths, sheet_names, active_cells = [], [], [], []
        excel = None
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
            for wb in excel.Workbooks:
                excel_files.append(wb.Name)
                file_paths.append(wb.FullName)
                try:
                    sht = wb.ActiveSheet
                    sheet_names.append(sht.Name)
                    cell_addr = sht.Application.ActiveCell.Address
                    active_cells.append(cell_addr)
                except Exception:
                    sheet_names.append("")
                    active_cells.append("")
        except Exception:
            pass
        finally:
            if excel is not None:
                del excel
            gc.collect()
            pythoncom.CoUninitialize()
        return excel_files, file_paths, sheet_names, active_cells

    def open_link_update_options(self):
        import tkinter as tk
        from tkinter import filedialog
        import os
        from tkinter import messagebox
    
        top = tk.Toplevel(self.root)
        top.title("Update Recent External Links Options")
        top.geometry("600x380")
        top.grab_set()
    
        frm = tk.Frame(top, padx=14, pady=12)
        frm.pack(fill="both", expand=True)
    
        default_font = ("Arial", 10, "bold")
    
        tk.Label(frm, text="Only update links modified within (days):", anchor="w", font=default_font).grid(row=0, column=0, sticky="w")
        entry_days = tk.Entry(frm, width=6, font=default_font)
        entry_days.insert(0, "14")
        entry_days.grid(row=0, column=1, sticky="w", padx=(6,0))
    
        cutoff_msg_var = tk.StringVar()
        def update_cutoff_msg(*args):
            try:
                days = int(entry_days.get())
                from datetime import datetime, timedelta
                threshold_date = datetime.now() - timedelta(days=days)
                cutoff_msg_var.set(f"Any external link file with last modified time on or after {threshold_date.strftime('%d-%m-%y %H:%M:%S')} will be refreshed and updated.")
            except Exception:
                cutoff_msg_var.set("")
        entry_days.bind("<KeyRelease>", update_cutoff_msg)
        update_cutoff_msg()
        tk.Label(frm, textvariable=cutoff_msg_var, fg="#0055aa", anchor="w", wraplength=440, justify="left", font=default_font).grid(row=1, column=0, columnspan=3, sticky="w", pady=(1, 8))
    
        show_link_var = tk.BooleanVar(value=True)
        show_full_path_var = tk.BooleanVar(value=True)
        show_last_modified_var = tk.BooleanVar(value=True)
        show_status_var = tk.BooleanVar(value=True)
        save_log_var = tk.BooleanVar(value=True)
        save_scan_summary_var = tk.BooleanVar(value=False)
    
        def on_full_path_check(*args):
            if show_full_path_var.get():
                if not show_link_var.get():
                    show_link_var.set(True)
        show_full_path_var.trace_add("write", on_full_path_check)

        
        tk.Checkbutton(frm, text="Show external link file name", variable=show_link_var, font=default_font).grid(row=2, column=0, columnspan=3, sticky="w")
        tk.Checkbutton(frm, text="Show external link source full path", variable=show_full_path_var, font=default_font).grid(row=3, column=0, columnspan=3, sticky="w")
        tk.Checkbutton(frm, text="Show last modified date", variable=show_last_modified_var, font=default_font).grid(row=4, column=0, columnspan=3, sticky="w")
        tk.Checkbutton(frm, text="Show update status", variable=show_status_var, font=default_font).grid(row=5, column=0, columnspan=3, sticky="w")
        tk.Checkbutton(frm, text="Save log file", variable=save_log_var, font=default_font).grid(row=6, column=0, columnspan=3, sticky="w")
    
        tk.Label(frm, text="Log file save to:", anchor="w", font=default_font).grid(row=7, column=0, sticky="w", pady=(6,0))
        entry_logdir = tk.Entry(frm, width=32, font=default_font)
        entry_logdir.insert(0, getattr(self, 'last_log_dir', r"D:\Pzone\Log"))
        entry_logdir.grid(row=7, column=1, sticky="w", pady=(6,0))
        def browse_logdir():
            d = filedialog.askdirectory(parent=top, initialdir=entry_logdir.get())
            if d:
                entry_logdir.delete(0, tk.END)
                entry_logdir.insert(0, d)
                self.last_log_dir = d
        browse_btn = tk.Button(frm, text="...", width=3, command=browse_logdir, font=default_font)
        browse_btn.grid(row=7, column=2, padx=(6,0), sticky="w", pady=(6,0))
    
        tk.Checkbutton(frm, text="Save external link scan summary", variable=save_scan_summary_var, font=default_font).grid(row=8, column=0, columnspan=3, sticky="w")
    
        tk.Label(frm, text="Scan summary save to:", anchor="w", font=default_font).grid(row=9, column=0, sticky="w", pady=(6,0))
        entry_summarydir = tk.Entry(frm, width=32, font=default_font)
        entry_summarydir.insert(0, getattr(self, 'last_summary_dir', r"D:\Pzone\Log"))
        entry_summarydir.grid(row=9, column=1, sticky="w", pady=(6,0))
        def browse_summarydir():
            d = filedialog.askdirectory(parent=top, initialdir=entry_summarydir.get())
            if d:
                entry_summarydir.delete(0, tk.END)
                entry_summarydir.insert(0, d)
                self.last_summary_dir = d
        browse_summary_btn = tk.Button(frm, text="...", width=3, command=browse_summarydir, font=default_font)
        browse_summary_btn.grid(row=9, column=2, padx=(6,0), sticky="w", pady=(6,0))
    
        btn_frame = tk.Frame(frm)
        btn_frame.grid(row=99, column=0, columnspan=3, pady=(18,0), sticky="w")
        ok_btn = tk.Button(btn_frame, text="OK", width=12, font=default_font)
        cancel_btn = tk.Button(btn_frame, text="Cancel", width=12, command=top.destroy, font=default_font)
        ok_btn.pack(side=tk.LEFT, padx=(0,14))
        cancel_btn.pack(side=tk.LEFT)
        
        def on_confirm():
            log_dir = entry_logdir.get()
            summary_dir = entry_summarydir.get()
            if not os.path.isdir(log_dir):
                messagebox.showerror("Error", "The selected log file directory does not exist. Please choose another directory.")
                return
            if save_scan_summary_var.get() and not os.path.isdir(summary_dir):
                messagebox.showerror("Error", "The selected scan summary directory does not exist. Please choose another directory.")
                return
            self.last_log_dir = log_dir
            self.last_summary_dir = summary_dir
            options = {
                "CHECK_DAYS": entry_days.get(),
                "SHOW_FULL_PATH": show_full_path_var.get(),
                "SHOW_LINK": show_link_var.get(),
                "SHOW_LAST_MODIFIED": show_last_modified_var.get(),
                "SHOW_STATUS": show_status_var.get(),
                "LOG_DIR": log_dir,
                "SAVE_LOG": save_log_var.get(),
                "SAVE_SCAN_SUMMARY": save_scan_summary_var.get(),
                "SUMMARY_DIR": summary_dir
            }
            popup = ConsolePopup(self.root, title="Update Recent External Links Console")
    
            def print_to_popup(msg):
                self.root.after(0, lambda: popup.print(msg))
            import threading
            threading.Thread(target=lambda: run_excel_link_update(options, print_func=print_to_popup), daemon=True).start()
            top.destroy()
        ok_btn.config(command=on_confirm)
    
        def on_confirm():
            log_dir = entry_logdir.get()
            summary_dir = entry_summarydir.get()
            if save_log_var.get() and not os.path.isdir(log_dir):
                messagebox.showerror("Error", "The selected log file directory does not exist. Please choose another directory.")
                return
            if save_scan_summary_var.get() and not os.path.isdir(summary_dir):
                messagebox.showerror("Error", "The selected scan summary directory does not exist. Please choose another directory.")
                return
            self.last_log_dir = log_dir
            self.last_summary_dir = summary_dir
            options = {
                "CHECK_DAYS": entry_days.get(),
                "SHOW_FULL_PATH": show_full_path_var.get(),
                "SHOW_LINK": show_link_var.get(),
                "SHOW_LAST_MODIFIED": show_last_modified_var.get(),
                "SHOW_STATUS": show_status_var.get(),
                "LOG_DIR": log_dir,
                "SAVE_LOG": save_log_var.get(),
                "SAVE_SCAN_SUMMARY": save_scan_summary_var.get(),
                "SUMMARY_DIR": summary_dir
            }
            popup = ConsolePopup(self.root, title="Update Recent External Links Console")
    
            def print_to_popup(msg):
                self.root.after(0, lambda: popup.print(msg))
            import threading
            threading.Thread(target=lambda: run_excel_link_update(options, print_func=print_to_popup), daemon=True).start()
            top.destroy()
        ok_btn.config(command=on_confirm)
    
    def treeview_sort(self, col):
        items = [self.tree.item(iid, "values") for iid in self.tree.get_children()]
        if not self.original_data or len(self.original_data) != len(items):
            self.original_data = list(items)
        if self.sort_states[col] == "none": self.sort_states[col] = "asc"
        elif self.sort_states[col] == "asc": self.sort_states[col] = "desc"
        else: self.sort_states[col] = "none"
        for other_col in self.sort_states:
            if other_col != col: self.sort_states[other_col] = "none"
        if self.sort_states[col] == "none":
            sorted_items = self.original_data
        else:
            reverse = self.sort_states[col] == "desc"
            if col == "col1": key_func = lambda x: (x[0] or "").lower()
            elif col == "col2": key_func = lambda x: (parse_mtime(x[1]) or datetime.min)
            else: key_func = lambda x: x
            sorted_items = sorted(items, key=key_func, reverse=reverse)
        for iid in self.tree.get_children(): self.tree.delete(iid)
        for v in sorted_items: self.tree.insert("", "end", values=v)
        self.update_treeview_font()

    def show_names(self):
        def update_gui(data):
            self.file_names, self.file_paths, self.sheet_names, self.active_cells = data
            self.count_label.config(text=f"({len(self.file_names)} files open)")
            for i in self.tree.get_children(): self.tree.delete(i)
            self.tree.heading("col1", text="File Path" if self.showing_path else "File Name")
            self.tree.heading("col2", text="Last Modified")
            if not self.file_names:
                self.tree.insert("", "end", values=("No Excel files are currently open.", ""))
            elif self.showing_path:
                for path in self.file_paths:
                    mtime = get_file_mtime_str(path)
                    self.tree.insert("", "end", values=(path, mtime))
            else:
                for i, name in enumerate(self.file_names):
                    mtime = get_file_mtime_str(self.file_paths[i])
                    self.tree.insert("", "end", values=(name, mtime))
            self.refresh_btn.config(state=tk.NORMAL)
            self.update_treeview_font()
            self.original_data.clear()
            self.sort_states = {"col1": "none", "col2": "none"}
            self.tree.xview_moveto(1.0)
        def scan_in_thread():
            kill_zombie_excel_processes()
            scan_data = self.get_open_excel_files()
            self.root.after(0, lambda: update_gui(scan_data))
        self.refresh_btn.config(state=tk.DISABLED)
        threading.Thread(target=scan_in_thread, daemon=True).start()

    def toggle_path(self):
        self.showing_path = not self.showing_path
        self.show_names()
        self.show_path_btn.config(text="Hide\nFile Path" if self.showing_path else "Show\nFile Path")

    def get_selected_workbooks(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("Notice", "Please select one or more Excel files before proceeding.")
            return []
        selected_workbooks = []
        current_shown_files = self.file_paths if self.showing_path else self.file_names
        for item_id in selected_items:
            item_values = self.tree.item(item_id, "values")
            if not item_values or not item_values[0]: continue
            selected_value = item_values[0]
            try:
                idx = current_shown_files.index(selected_value)
                if 0 <= idx < len(self.file_names):
                    selected_workbooks.append((self.file_names[idx], self.file_paths[idx], self.sheet_names[idx], self.active_cells[idx]))
            except ValueError:
                continue
        return selected_workbooks

    def on_selection_change(self, event):
        selected_items = self.tree.selection()
        captions = []
        for item_id in selected_items:
            item_values = self.tree.item(item_id, "values")
            if not item_values or not item_values[0]: continue
            selected_value = item_values[0]
            if self.showing_path:
                captions.append(os.path.basename(selected_value))
            else:
                captions.append(selected_value)
        self.target_captions = captions
        self.select_all_var.set(
            len(self.tree.get_children()) > 0 and
            len(self.tree.selection()) == len(self.tree.get_children())
        )

    def on_treeview_heading_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region == "heading":
            col = self.tree.identify_column(event.x)
            if col == "#1": self.treeview_sort("col1")
            elif col == "#2": self.treeview_sort("col2")

    def activate_selected_workbooks(self):
        self.on_selection_change(None)
        if not self.target_captions:
            messagebox.showinfo("Notice", "Please select one or more Excel files to activate.")
            return
        offset_x, offset_y, start_x, start_y = 40, 40, 100, 100
        activated_hwnds, window_index = set(), 0
        def enum_handler(hwnd, ctx):
            nonlocal window_index
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                for caption in ctx["captions"]:
                    if caption in title and " - Excel" in title and hwnd not in activated_hwnds:
                        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                        win32gui.SetForegroundWindow(hwnd)
                        x, y = start_x + window_index * offset_x, start_y + window_index * offset_y
                        rect = win32gui.GetWindowRect(hwnd)
                        w, h = rect[2] - rect[0], rect[3] - rect[1]
                        win32gui.SetWindowPos(hwnd, None, x, y, w, h, win32con.SWP_NOZORDER | win32con.SWP_SHOWWINDOW)
                        activated_hwnds.add(hwnd)
                        window_index += 1
                        time.sleep(0.05)
        win32gui.EnumWindows(enum_handler, {"captions": self.target_captions})

    def save_selected_workbooks(self):
        selected = self.get_selected_workbooks()
        if not selected: return
        # your patch added between the existing code
        show_console = self.show_console_progress_var.get() if hasattr(self, "show_console_progress_var") else True
        popup = ConsolePopup(self.root, title="Save Selected Progress") if show_console else None
        def print_to_popup(msg):
            if popup: self.root.after(0, lambda: popup.print(msg))
        def thread_job():
            pythoncom.CoInitialize()
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
                print_to_popup(f"Saving {len(selected)} file(s)...")
                print_to_popup("-" * 80)
                for idx, (name, path, _, _) in enumerate(selected, 1):
                    print_to_popup(f"({idx}/{len(selected)}) Saving: {path}")
                    t0 = time.time()
                    saved = False
                    for wb in excel.Workbooks:
                        if wb.Name == name and wb.FullName == path:
                            try:
                                wb.Save()
                                print_to_popup(f"({idx}/{len(selected)}) Saved: {path}")
                                saved = True
                            except Exception as e:
                                print_to_popup(f"({idx}/{len(selected)}) Save failed: {path} ({e})")
                            break
                    if not saved:
                        print_to_popup(f"({idx}/{len(selected)}) Workbook not found in Excel: {path}")
                    t1 = time.time()
                    used_sec = t1 - t0
                    print_to_popup(f"used time: {used_sec:.2f} sec")
                    print_to_popup("-" * 80)
                print_to_popup("Save selected complete.")
                self.root.after(0, lambda: messagebox.showinfo("Complete", "Selected Excel files have been saved successfully."))
            except Exception as e:
                print_to_popup(f"Error: {str(e)}")
                self.root.after(0, lambda e=e: messagebox.showerror("Error", f"An error occurred while saving the files:\n{str(e)}"))
            finally:
                pythoncom.CoUninitialize()
                self.root.after(0, self.show_names)
        threading.Thread(target=thread_job, daemon=True).start()


    def close_selected_workbooks(self, save_before_close=False):
        selected = self.get_selected_workbooks()
        if not selected: return
        # your patch added between the existing code
        show_console = self.show_console_progress_var.get() if hasattr(self, "show_console_progress_var") else True
        popup = ConsolePopup(self.root, title="Save and Close Progress" if save_before_close else "Close Progress") if show_console else None
        def print_to_popup(msg):
            if popup: self.root.after(0, lambda: popup.print(msg))
        def thread_job():
            pythoncom.CoInitialize()
            excel = None
            orig_alert = None
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
                try:
                    orig_alert = excel.DisplayAlerts
                    excel.DisplayAlerts = False
                except Exception:
                    pass
                print_to_popup(f"{'Save and close' if save_before_close else 'Close'} {len(selected)} file(s)...")
                print_to_popup("-" * 80)
                for idx, (name, path, _, _) in enumerate(selected, 1):
                    t0 = time.time()
                    closed = False
                    if save_before_close:
                        print_to_popup(f"({idx}/{len(selected)}) Saving and closing: {path}")
                    else:
                        print_to_popup(f"({idx}/{len(selected)}) Closing: {path}")
                    for wb in excel.Workbooks:
                        if wb.Name == name and wb.FullName == path:
                            try:
                                wb.Close(SaveChanges=save_before_close)
                                closed = True
                                if save_before_close:
                                    print_to_popup(f"({idx}/{len(selected)}) Saved and closed: {path}")
                                else:
                                    print_to_popup(f"({idx}/{len(selected)}) Closed: {path}")
                            except Exception as e:
                                print_to_popup(f"({idx}/{len(selected)}) Close failed: {path} ({e})")
                            break
                    if not closed:
                        print_to_popup(f"({idx}/{len(selected)}) Workbook not found in Excel: {path}")
                    t1 = time.time()
                    used_sec = t1 - t0
                    print_to_popup(f"used time: {used_sec:.2f} sec")
                    print_to_popup("-" * 80)
                if excel.Workbooks.Count == 0:
                    excel.Quit()
                print_to_popup(f"{'Save and close' if save_before_close else 'Close'} selected complete.")
                self.root.after(0, lambda: messagebox.showinfo("Complete", f"Selected files have been {'saved and ' if save_before_close else ''}closed."))
            except Exception as e:
                print_to_popup(f"Error: {str(e)}")
                self.root.after(0, lambda e=e: messagebox.showerror("Error", f"An error occurred while closing the files:\n{str(e)}"))
            finally:
                if excel is not None:
                    try:
                        if hasattr(excel, "DisplayAlerts") and orig_alert is not None:
                            excel.DisplayAlerts = orig_alert
                    except Exception:
                        pass
                    del excel
                gc.collect()
                pythoncom.CoUninitialize()
                self.root.after(0, self.show_names)
        threading.Thread(target=thread_job, daemon=True).start()

    def minimize_all_excel(self):
        def enum_handler(hwnd, ctx):
            if win32gui.IsWindowVisible(hwnd) and " - Excel" in win32gui.GetWindowText(hwnd):
                win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)
        win32gui.EnumWindows(enum_handler, None)

    def save_session(self):
        selected = self.get_selected_workbooks()
        if not selected: return
        file_path = filedialog.asksaveasfilename(
            title="Save Session", defaultextension=".xlsx",
            filetypes=[("Excel Session", "*.xlsx"), ("All Files", "*.*")]
        )
        if not file_path: return
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        base, ext = os.path.splitext(file_path)
        file_path_with_ts = f"{base}_{timestamp}{ext}"
        pythoncom.CoInitialize()
        try:
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
                excel.Quit()
            except Exception:
                pass
        finally:
            pythoncom.CoUninitialize()
        time.sleep(1)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Session"
        ws.append(["File Path", "Sheet Name", "Cell Address"])
        for _, path, sheet, cell in selected: ws.append([path, sheet, cell])
        wb.save(file_path_with_ts)
        messagebox.showinfo("Success", f"Session saved at:\n{file_path_with_ts}")
        self.load_session_from_path(file_path_with_ts)

    def load_session(self):
        current_files, _, _, _ = self.get_open_excel_files()
        if current_files:
            messagebox.showwarning("Warning", "Please close all currently open Excel files before loading a session.")
            self.show_names()
            return
        file_path = filedialog.askopenfilename(
            title="Load Session", filetypes=[("Excel Session", "*.xlsx"), ("All Files", "*.*")]
        )
        if not file_path or not os.path.exists(file_path): return
        # your patch added between the existing code
        show_console = self.show_console_progress_var.get() if hasattr(self, "show_console_progress_var") else True
        popup = ConsolePopup(self.root, title="Load Session Progress") if show_console else None
        def print_to_popup(msg):
            if popup: self.root.after(0, lambda: popup.print(msg))
        def thread_job():
            pythoncom.CoInitialize()
            excel = None
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                valid_rows = [r for r in rows if r and r[0] and os.path.exists(r[0])]
                if not valid_rows:
                    print_to_popup("No valid file paths found.")
                    self.root.after(0, lambda: messagebox.showwarning("Warning", "No valid file paths found."))
                    return
                print_to_popup(f"Loading session from {file_path} ({len(valid_rows)} file(s))")
                print_to_popup("-" * 80)
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True
                excel.AskToUpdateLinks = False
                for idx, r in enumerate(valid_rows, 1):
                    path, sheet, cell = (r[0], r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)
                    print_to_popup(f"({idx}/{len(valid_rows)}) Opening: {path}")
                    t0 = time.time()
                    try:
                        wb_xl = excel.Workbooks.Open(Filename=path, UpdateLinks=0)
                        try:
                            excel.Visible = True
                        except Exception:
                            pass
                        if sheet:
                            try:
                                sht = wb_xl.Sheets(sheet)
                                sht.Activate()
                                if cell: sht.Range(cell).Select()
                            except Exception as e:
                                print_to_popup(f"  (Sheet/Cell select error: {e})")
                        print_to_popup(f"({idx}/{len(valid_rows)}) Opened: {path}")
                    except Exception as e:
                        print_to_popup(f"({idx}/{len(valid_rows)}) Failed to open: {path} ({e})")
                    t1 = time.time()
                    used_sec = t1 - t0
                    print_to_popup(f"used time: {used_sec:.2f} sec")
                    print_to_popup("-" * 80)
                excel.AskToUpdateLinks = True
                print_to_popup(f"All files loaded. Total: {len(valid_rows)}")
                self.root.after(0, lambda: messagebox.showinfo("Complete", f"{len(valid_rows)} file(s) opened."))
            except Exception as e:
                print_to_popup(f"Error loading session: {str(e)}")
                self.root.after(0, lambda e=e: messagebox.showerror("Error", f"Error loading session:\n{str(e)}"))
            finally:
                if excel is not None:
                    del excel
                gc.collect()
                pythoncom.CoUninitialize()
                self.root.after(200, self.show_names)
        threading.Thread(target=thread_job, daemon=True).start()

def load_session_from_path(self, file_path):
    def thread_job():
        pythoncom.CoInitialize()
        excel = None
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            valid_rows = [r for r in rows if r and r[0] and os.path.exists(r[0])]
            if not valid_rows:
                self.root.after(0, lambda: messagebox.showwarning("Warning", "No valid file paths found."))
                return
            excel = win32com.client.Dispatch("Excel.Application")
            try:
                excel.Visible = True
            except Exception:
                pass
            excel.AskToUpdateLinks = False
            for r in valid_rows:
                path, sheet, cell = (r[0], r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)
                try:
                    wb_xl = excel.Workbooks.Open(Filename=path, UpdateLinks=0)
                    try:
                        excel.Visible = True
                    except Exception:
                        pass
                    if sheet:
                        try:
                            sht = wb_xl.Sheets(sheet)
                            sht.Activate()
                            if cell: sht.Range(cell).Select()
                        except Exception: pass
                except Exception: pass
            excel.AskToUpdateLinks = True
            self.root.after(0, lambda: messagebox.showinfo("Complete", f"{len(valid_rows)} file(s) opened."))
        except Exception as e:
            self.root.after(0, lambda e=e: messagebox.showerror("Error", f"Error loading session:\n{str(e)}"))
        finally:
            if excel is not None:
                del excel
            gc.collect()
            pythoncom.CoUninitialize()
            self.root.after(200, self.show_names)
    threading.Thread(target=thread_job, daemon=True).start()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelSessionManagerApp(root)
    root.mainloop()
