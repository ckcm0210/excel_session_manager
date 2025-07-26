"""
Session Manager for Excel Session Manager

This module handles saving and loading Excel session files,
including file validation and session data management.
"""

import os
import openpyxl
from datetime import datetime
from tkinter import filedialog, messagebox
import threading
import pythoncom
import win32com.client
import time
import gc
from ui.console_popup import ConsolePopup
from ui.dialogs.file_selector import FileSelectionDialog
from core.performance_monitor import timed_operation


class SessionManager:
    """
    Manages Excel session saving and loading operations.
    
    Handles creating session files with workbook information,
    loading sessions with file selection, and validating session data.
    """
    
    def __init__(self, parent_window):
        """
        Initialize the session manager.
        
        Args:
            parent_window: Parent window for dialogs
        """
        self.parent = parent_window
        
    def save_session(self, selected_workbooks):
        """
        Save selected workbooks to a session file.
        
        Args:
            selected_workbooks: List of tuples (name, path, sheet, cell)
            
        Returns:
            str: Path to saved session file, or None if cancelled
        """
        if not selected_workbooks:
            return None
        
        # Add performance monitoring with workbook count
        from core.performance_monitor import get_performance_monitor
        monitor = get_performance_monitor()
        op_id = monitor.start_operation("save_session", {'workbook_count': len(selected_workbooks)})
        
        try:
            result = self._save_session_impl(selected_workbooks)
            monitor.end_operation(op_id, success=True)
            return result
        except Exception as e:
            monitor.end_operation(op_id, success=False)
            raise
    
    def _save_session_impl(self, selected_workbooks):
        """Implementation of save session."""
        file_path = filedialog.asksaveasfilename(
            title="Save Session", 
            defaultextension=".xlsx",
            filetypes=[("Excel Session", "*.xlsx"), ("All Files", "*.*")]
        )
        if not file_path:
            return None
            
        # Add timestamp to filename
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        base, ext = os.path.splitext(file_path)
        file_path_with_ts = f"{base}_{timestamp}{ext}"

        try:
            # Create session workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Session"
            ws.append(["File Path", "Sheet Name", "Cell Address"])
            
            # Add workbook data
            for _, path, sheet, cell in selected_workbooks:
                ws.append([path, sheet, cell])
                
            wb.save(file_path_with_ts)
            messagebox.showinfo("Success", f"Session saved at:\n{file_path_with_ts}")
            return file_path_with_ts
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save session:\n{str(e)}")
            return None
    
    def load_session(self, get_open_files_func, show_console_var=None):
        """
        Load a session file and open selected Excel files.
        
        Args:
            get_open_files_func: Function to get currently open Excel files
            show_console_var: BooleanVar for showing progress console
        """
        # Check if Excel files are already open
        current_files, _, _, _ = get_open_files_func()
        if current_files:
            messagebox.showwarning(
                "Warning", 
                "Please close all currently open Excel files before loading a session."
            )
            return
            
        # Select session file
        file_path = filedialog.askopenfilename(
            title="Load Session", 
            filetypes=[("Excel Session", "*.xlsx"), ("All Files", "*.*")]
        )
        if not file_path or not os.path.exists(file_path):
            return
            
        # Load and validate session file
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            all_file_paths = [r[0] for r in rows if r and r[0]]
            valid_file_paths = [path for path in all_file_paths if os.path.exists(path)]
            
            if not valid_file_paths:
                messagebox.showwarning("Warning", "No valid file paths found in session.")
                return
                
            # Show file selection dialog
            dialog = FileSelectionDialog(self.parent, valid_file_paths, "Select Files to Load from Session")
            result, selected_files = dialog.show()
            
            if result != "ok" or not selected_files:
                return
            
            # Add performance monitoring with file count
            from core.performance_monitor import get_performance_monitor
            monitor = get_performance_monitor()
            op_id = monitor.start_operation("load_session", {'selected_count': len(selected_files)})
                
            # Create filtered rows for selected files only
            selected_rows = []
            for r in rows:
                if r and r[0] and r[0] in selected_files:
                    selected_rows.append(r)
                    
        except Exception as e:
            messagebox.showerror("Error", f"Error reading session file:\n{str(e)}")
            return
            
        # Load files with progress console
        show_console = show_console_var.get() if show_console_var else True
        popup = ConsolePopup(self.parent, title="Load Session Progress") if show_console else None
        
        def print_to_popup(msg):
            if popup:
                self.parent.after(0, lambda: popup.print(msg))
                
        def thread_job():
            try:
                self._load_files_thread(selected_rows, print_to_popup)
                monitor.end_operation(op_id, success=True)
            except Exception as e:
                monitor.end_operation(op_id, success=False)
                raise
            
        threading.Thread(target=thread_job, daemon=True).start()
    
    def _load_files_thread(self, selected_rows, print_func):
        """
        Thread function to load Excel files from session.
        
        Args:
            selected_rows: List of session data rows
            print_func: Function to print progress messages
        """
        pythoncom.CoInitialize()
        excel = None
        
        try:
            if not selected_rows:
                print_func("No files selected to load.")
                self.parent.after(0, lambda: messagebox.showwarning("Warning", "No files selected to load."))
                return
                
            print_func(f"Loading selected files from session ({len(selected_rows)} file(s))")
            print_func("-" * 80)
            
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True
            excel.AskToUpdateLinks = False
            
            for idx, r in enumerate(selected_rows, 1):
                path, sheet, cell = (r[0], r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                print_func(f"{ts} | ({idx}/{len(selected_rows)}) Opening: {path}")
                t0 = time.time()
                
                try:
                    wb_xl = excel.Workbooks.Open(Filename=path, UpdateLinks=0)
                    if wb_xl.ReadOnly:
                        ts2 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        print_func(f'{ts2} |     File "{path}" is opened in read-only mode. Changes may not be saved.')
                    
                    try:
                        excel.Visible = True
                    except Exception:
                        pass
                        
                    if sheet:
                        try:
                            sht = wb_xl.Sheets(sheet)
                            sht.Activate()
                            if cell:
                                sht.Range(cell).Select()
                        except Exception as e:
                            print_func(f"  (Sheet/Cell select error: {e})")
                            
                    ts3 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    print_func(f"{ts3} | ({idx}/{len(selected_rows)}) Opened: {path}")
                    
                except Exception as e:
                    ts4 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    print_func(f"{ts4} | ({idx}/{len(selected_rows)}) Failed to open: {path} ({e})")
                    
                t1 = time.time()
                used_sec = t1 - t0
                print_func(f"used time: {used_sec:.2f} sec")
                print_func("-" * 80)
                
            excel.AskToUpdateLinks = True
            print_func(f"All files loaded. Total: {len(selected_rows)}")
            self.parent.after(0, lambda: messagebox.showinfo("Complete", f"{len(selected_rows)} file(s) opened."))
            
        except Exception as e:
            print_func(f"Error loading session: {str(e)}")
            self.parent.after(0, lambda e=e: messagebox.showerror("Error", f"Error loading session:\n{str(e)}"))
            
        finally:
            if excel is not None:
                del excel
            gc.collect()
            pythoncom.CoUninitialize()