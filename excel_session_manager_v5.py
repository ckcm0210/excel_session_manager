import tkinter as tk
from tkinter import ttk, messagebox, filedialog, font
import pythoncom
import win32com.client
import win32gui
import win32con
import time
import openpyxl
import os
from datetime import datetime
import threading
from PIL import Image, ImageTk  # 記得要 pip install pillow

# ==============================================================================
# (所有輔助 Class 和 Function 保持不變)
# ==============================================================================
MONO_FONTS = [
    "Consolas", "Courier New", "Fira Code", "JetBrains Mono",
    "DejaVu Sans Mono", "Source Code Pro", "Monaco"
]

def calc_row_height(fsize):
    return int(fsize * 2.1)

def calc_col2_width(fsize):
    return int(fsize * 11.5)

class DragSelectTreeview(ttk.Treeview):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._drag_start = None
        self._drag_mode = None
        self.bind("<Button-1>", self._on_click)
        self.bind("<B1-Motion>", self._on_drag)
        self.bind("<ButtonRelease-1>", self._on_release)

    def _on_click(self, event):
        iid = self.identify_row(event.y)
        if not iid:
            self.selection_remove(self.selection())
            return
        self.focus(iid)
        if iid in self.selection():
            self.selection_remove(iid)
            self._drag_mode = "unselect"
        else:
            self.selection_add(iid)
            self._drag_mode = "select"
        self._drag_start = iid
        return "break"

    def _on_drag(self, event):
        iid = self.identify_row(event.y)
        if not iid or not self._drag_start:
            return
        iids = self.get_children()
        start_idx = iids.index(self._drag_start)
        cur_idx = iids.index(iid)
        lo = min(start_idx, cur_idx)
        hi = max(start_idx, cur_idx)
        if self._drag_mode == "unselect":
            for i in iids[lo:hi+1]:
                self.selection_remove(i)
        else:
            for i in iids[lo:hi+1]:
                self.selection_add(i)
        return "break"

    def _on_release(self, event):
        self._drag_start = None
        self._drag_mode = None
        self.event_generate("<<TreeviewSelect>>")
        return "break"

def get_file_mtime_str(path):
    if os.path.exists(path):
        ts = os.path.getmtime(path)
        dt = datetime.fromtimestamp(ts)
        return f"{dt.day}/{dt.month}/{dt.year} {dt.hour}:{dt.minute:02d}"
    else:
        return ''

def parse_mtime(mtime_str):
    try:
        parts = mtime_str.split()
        date, time_part = parts[0], parts[1]
        day, month, year = map(int, date.split('/'))
        hour, minute = map(int, time_part.split(':'))
        return datetime(year, month, day, hour, minute)
    except Exception:
        return None

# ==============================================================================
# 主要應用程式 Class
# ==============================================================================
class ExcelSessionManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Session Manager")
        self.root.geometry("1200x620")
        self.root.minsize(600, 500)
        self.root.protocol("WM_DELETE_WINDOW", lambda: self.root.destroy())

        self.is_mini = False
        self.normal_geometry = "1200x620"
        self.mini_side = 180
        self.mini_geometry = f"{self.mini_side}x{self.mini_side}+150+150"
        self.floating_icon_btn = None

        self.setup_ui()
        self.show_names()

    def setup_ui(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill="both")

        frame_session = ttk.Frame(self.notebook)
        self.notebook.add(frame_session, text="Session Manager")

        font_frame = tk.Frame(frame_session)
        font_frame.pack(pady=(5, 0), padx=10, anchor="nw", fill='x')

        self.pin_btn = tk.Button(
            font_frame,
            text="Mini Widget (Floating Pin)",
            font=("Microsoft JhengHei", 10),
            command=self.enter_mini,
            relief="raised",
            borderwidth=1,
            padx=5,
            pady=2
        )
        self.pin_btn.pack(side=tk.RIGHT, padx=(10, 0))

        tk.Label(font_frame, text="Font:").pack(side=tk.LEFT)
        self.font_family_var = tk.StringVar(value=MONO_FONTS[0])
        self.font_size_var = tk.IntVar(value=12)
        font_family_menu = ttk.OptionMenu(font_frame, self.font_family_var, MONO_FONTS[0], *MONO_FONTS, command=self.update_treeview_font)
        font_family_menu.pack(side=tk.LEFT, padx=(5, 15))

        tk.Label(font_frame, text="Size:").pack(side=tk.LEFT)
        font_size_spin = ttk.Spinbox(font_frame, from_=8, to=32, textvariable=self.font_size_var, width=4, command=self.update_treeview_font)
        font_size_spin.pack(side=tk.LEFT)
        self.font_size_var.trace_add("write", lambda *a: self.update_treeview_font())

        title_frame = tk.Frame(frame_session)
        title_frame.pack(pady=(10, 0), padx=10, anchor="w", fill='x')

        main_label = tk.Label(title_frame, text="Active Excel Session:", font=("Arial", 16, "bold"))
        main_label.pack(side=tk.LEFT)

        self.count_label = tk.Label(title_frame, text="", font=("Arial", 12))
        self.count_label.pack(side=tk.LEFT, padx=(5, 0), pady=(2,0))

        main_frame = tk.Frame(frame_session)
        main_frame.pack(pady=(5, 10), padx=10, expand=True, fill='both')

        side_btn_frame = tk.Frame(main_frame, width=120)
        side_btn_frame.pack(side=tk.RIGHT, padx=(10, 0), pady=10, fill='y')
        side_btn_frame.pack_propagate(False)
        btn_props = {'width': 20, 'height': 2, 'wraplength': 140, 'font': ("Arial", 10, "bold")}

        self.refresh_btn = tk.Button(side_btn_frame, text="Refresh\nList", **btn_props, command=self.show_names)
        self.refresh_btn.pack(pady=5, anchor='n')

        self.show_path_btn = tk.Button(side_btn_frame, text="Show\nFile Path", **btn_props, command=self.toggle_path)
        self.show_path_btn.pack(pady=5, anchor='n')

        activate_btn = tk.Button(side_btn_frame, text="Activate\nSelected", **btn_props, command=self.activate_selected_workbooks)
        activate_btn.pack(pady=5, anchor='n')

        minimize_btn = tk.Button(side_btn_frame, text="Minimize All\nExcel", **btn_props, command=self.minimize_all_excel)
        minimize_btn.pack(pady=5, anchor='n')

        save_sess_btn = tk.Button(side_btn_frame, text="Save\nSession", **btn_props, command=self.save_session)
        save_sess_btn.pack(pady=5, anchor='n')

        load_sess_btn = tk.Button(side_btn_frame, text="Load\nSession", **btn_props, command=self.load_session)
        load_sess_btn.pack(pady=5, anchor='n')

        btn1 = tk.Button(side_btn_frame, text="Save\nSelected", **btn_props, command=self.save_selected_workbooks)
        btn1.pack(pady=5, anchor='n')

        btn2 = tk.Button(side_btn_frame, text="Close\nWithout Saving", **btn_props, command=lambda: self.close_selected_workbooks(False))
        btn2.pack(pady=5, anchor='n')

        btn3 = tk.Button(side_btn_frame, text="Save and Close\nSelected", **btn_props, command=lambda: self.close_selected_workbooks(True))
        btn3.pack(pady=5, anchor='n')

        listbox_frame = tk.Frame(main_frame)
        listbox_frame.pack(side=tk.LEFT, fill='both', expand=True, pady=10)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 12, "bold"))
        style.configure("Treeview", rowheight=calc_row_height(self.font_size_var.get()))

        self.tree = DragSelectTreeview(listbox_frame, columns=("col1", "col2"), show="headings", selectmode="extended")
        self.tree.column("col1", anchor="w")
        self.tree.column("col2", anchor="e", width=125, stretch=False)
        self.tree.pack(expand=True, fill='both', padx=8, pady=8)

        self.tree.bind('<<TreeviewSelect>>', self.on_selection_change)
        self.tree.bind("<Button-1>", self.on_treeview_heading_click, add="+")

        self.file_names, self.file_paths, self.sheet_names, self.active_cells = [], [], [], []
        self.showing_path = False
        self.target_captions = []
        self.sort_states = {"col1": "none", "col2": "none"}
        self.original_data = []

    # ==============================================================================
    # Mini Widget 邏輯 (已修正)
    # ==============================================================================
    def enter_mini(self):
        if self.is_mini:
            return
        self.is_mini = True
        self.normal_geometry = self.root.geometry()
        self.notebook.pack_forget()

        # ==============================================================
        # **修改點**: 暫時解除最小尺寸限制
        # ==============================================================
        self.root.minsize(self.mini_side, self.mini_side)
        self.root.geometry(self.mini_geometry)
        self.root.resizable(False, False)
        self.root.wm_attributes("-topmost", 1)

        icon_path = "maximize_full_screen.png"
        icon_size = (self.mini_side - 40, self.mini_side - 40)
        if os.path.exists(icon_path):
            try:
                pil_image = Image.open(icon_path).resize(icon_size, Image.LANCZOS)
                icon_image = ImageTk.PhotoImage(pil_image)
                self.floating_icon_btn = tk.Button(
                    self.root, image=icon_image, width=icon_size[0], height=icon_size[1],
                    command=self.exit_mini, borderwidth=0, relief="flat"
                )
                self.floating_icon_btn.image = icon_image
            except Exception:
                self.create_fallback_icon()
        else:
            self.create_fallback_icon()
        self.floating_icon_btn.pack(expand=True, padx=10, pady=10)

    def exit_mini(self):
        if not self.is_mini:
            return
        self.is_mini = False
        if self.floating_icon_btn:
            self.floating_icon_btn.destroy()
            self.floating_icon_btn = None

        # ==============================================================
        # **修改點**: 還原返原本嘅最小尺寸限制
        # ==============================================================
        self.root.minsize(600, 500)
        self.root.geometry(self.normal_geometry)
        self.root.resizable(True, True)
        self.root.wm_attributes("-topmost", 0)
        self.notebook.pack(expand=True, fill="both")

    def create_fallback_icon(self):
        self.floating_icon_btn = tk.Button(
            self.root, text="🗔", font=("Segoe UI Emoji", 48),
            command=self.exit_mini, borderwidth=0, relief="flat"
        )

    # ==============================================================================
    # (以下所有 method 保持不變)
    # ==============================================================================
    def update_treeview_font(self, *args):
        new_font = (self.font_family_var.get(), self.font_size_var.get())
        style = ttk.Style()
        style.configure("Treeview.Heading", font=(self.font_family_var.get(), self.font_size_var.get(), "bold"))
        self.tree.tag_configure('custom_font', font=new_font)
        for iid in self.tree.get_children():
            self.tree.item(iid, tags=('custom_font',))
        row_h = calc_row_height(self.font_size_var.get())
        style.configure("Treeview", rowheight=row_h)
        self.tree.column("col2", width=125, stretch=False)

    def get_open_excel_files(self):
        pythoncom.CoInitialize()
        excel_files, file_paths, sheet_names, active_cells = [], [], [], []
        excel = None
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
            for wb in excel.Workbooks:
                excel_files.append(wb.Name)
                file_paths.append(wb.FullName)
                try:
                    sht = wb.ActiveSheet
                    sheet_names.append(sht.Name)
                    cell_addr = sht.Application.ActiveCell.Address
                    active_cells.append(cell_addr)
                except Exception:
                    sheet_names.append("")
                    active_cells.append("")
        except Exception:
            pass
        finally:
            excel = None
            pythoncom.CoUninitialize()
        return excel_files, file_paths, sheet_names, active_cells

    def treeview_sort(self, col):
        items = [self.tree.item(iid, "values") for iid in self.tree.get_children()]
        if not self.original_data or len(self.original_data) != len(items):
            self.original_data = list(items)
        if self.sort_states[col] == "none": self.sort_states[col] = "asc"
        elif self.sort_states[col] == "asc": self.sort_states[col] = "desc"
        else: self.sort_states[col] = "none"
        for other_col in self.sort_states:
            if other_col != col: self.sort_states[other_col] = "none"
        if self.sort_states[col] == "none":
            sorted_items = self.original_data
        else:
            reverse = self.sort_states[col] == "desc"
            if col == "col1": key_func = lambda x: (x[0] or "").lower()
            elif col == "col2": key_func = lambda x: (parse_mtime(x[1]) or datetime.min)
            else: key_func = lambda x: x
            sorted_items = sorted(items, key=key_func, reverse=reverse)
        for iid in self.tree.get_children(): self.tree.delete(iid)
        for v in sorted_items: self.tree.insert("", "end", values=v)
        self.update_treeview_font()

    def show_names(self):
        def update_gui(data):
            self.file_names, self.file_paths, self.sheet_names, self.active_cells = data
            self.count_label.config(text=f"({len(self.file_names)} files open)")
            for i in self.tree.get_children(): self.tree.delete(i)
            self.tree.heading("col1", text="File Path" if self.showing_path else "File Name")
            self.tree.heading("col2", text="Last Modified")
            if not self.file_names:
                self.tree.insert("", "end", values=("No Excel files are currently open.", ""))
            elif self.showing_path:
                for path in self.file_paths:
                    mtime = get_file_mtime_str(path)
                    self.tree.insert("", "end", values=(path, mtime))
            else:
                for i, name in enumerate(self.file_names):
                    mtime = get_file_mtime_str(self.file_paths[i])
                    self.tree.insert("", "end", values=(name, mtime))
            self.refresh_btn.config(state=tk.NORMAL)
            self.update_treeview_font()
            self.original_data.clear()
            self.sort_states = {"col1": "none", "col2": "none"}
        def scan_in_thread():
            scan_data = self.get_open_excel_files()
            self.root.after(0, lambda: update_gui(scan_data))
        self.refresh_btn.config(state=tk.DISABLED)
        threading.Thread(target=scan_in_thread, daemon=True).start()

    def toggle_path(self):
        self.showing_path = not self.showing_path
        self.show_names()
        self.show_path_btn.config(text="Hide\nFile Path" if self.showing_path else "Show\nFile Path")

    def get_selected_workbooks(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("Notice", "Please select one or more Excel files before proceeding.")
            return []
        selected_workbooks = []
        current_shown_files = self.file_paths if self.showing_path else self.file_names
        for item_id in selected_items:
            item_values = self.tree.item(item_id, "values")
            if not item_values or not item_values[0]: continue
            selected_value = item_values[0]
            try:
                idx = current_shown_files.index(selected_value)
                if 0 <= idx < len(self.file_names):
                    selected_workbooks.append((self.file_names[idx], self.file_paths[idx], self.sheet_names[idx], self.active_cells[idx]))
            except ValueError:
                continue
        return selected_workbooks

    def on_selection_change(self, event):
        selected_items = self.tree.selection()
        captions = []
        for item_id in selected_items:
            item_values = self.tree.item(item_id, "values")
            if not item_values or not item_values[0]: continue
            selected_value = item_values[0]
            if self.showing_path:
                captions.append(os.path.basename(selected_value))
            else:
                captions.append(selected_value)
        self.target_captions = captions

    def on_treeview_heading_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region == "heading":
            col = self.tree.identify_column(event.x)
            if col == "#1": self.treeview_sort("col1")
            elif col == "#2": self.treeview_sort("col2")

    def activate_selected_workbooks(self):
        self.on_selection_change(None)
        if not self.target_captions:
            messagebox.showinfo("Notice", "Please select one or more Excel files to activate.")
            return
        offset_x, offset_y, start_x, start_y = 40, 40, 100, 100
        activated_hwnds, window_index = set(), 0
        def enum_handler(hwnd, ctx):
            nonlocal window_index
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                for caption in ctx["captions"]:
                    if caption in title and " - Excel" in title and hwnd not in activated_hwnds:
                        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                        win32gui.SetForegroundWindow(hwnd)
                        x, y = start_x + window_index * offset_x, start_y + window_index * offset_y
                        rect = win32gui.GetWindowRect(hwnd)
                        w, h = rect[2] - rect[0], rect[3] - rect[1]
                        win32gui.SetWindowPos(hwnd, None, x, y, w, h, win32con.SWP_NOZORDER | win32con.SWP_SHOWWINDOW)
                        activated_hwnds.add(hwnd)
                        window_index += 1
                        time.sleep(0.05)
        win32gui.EnumWindows(enum_handler, {"captions": self.target_captions})

    def save_selected_workbooks(self):
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
            selected = self.get_selected_workbooks()
            if not selected: return
            for name, path, _, _ in selected:
                for wb in excel.Workbooks:
                    if wb.Name == name and wb.FullName == path: wb.Save()
            messagebox.showinfo("Complete", "Selected Excel files have been saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving the files:\n{str(e)}")
        finally:
            pythoncom.CoUninitialize()
            self.show_names()

    def close_selected_workbooks(self, save_before_close=False):
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
            selected = self.get_selected_workbooks()
            if not selected: return
            for name, path, _, _ in selected:
                for wb in excel.Workbooks:
                    if wb.Name == name and wb.FullName == path: wb.Close(SaveChanges=save_before_close)
            messagebox.showinfo("Complete", f"Selected files have been {'saved and ' if save_before_close else ''}closed.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while closing the files:\n{str(e)}")
        finally:
            pythoncom.CoUninitialize()
            self.show_names()

    def minimize_all_excel(self):
        def enum_handler(hwnd, ctx):
            if win32gui.IsWindowVisible(hwnd) and " - Excel" in win32gui.GetWindowText(hwnd):
                win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)
        win32gui.EnumWindows(enum_handler, None)

    def save_session(self):
        selected = self.get_selected_workbooks()
        if not selected: return
        file_path = filedialog.asksaveasfilename(
            title="Save Session", defaultextension=".xlsx",
            filetypes=[("Excel Session", "*.xlsx"), ("All Files", "*.*")]
        )
        if not file_path: return
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        base, ext = os.path.splitext(file_path)
        file_path_with_ts = f"{base}_{timestamp}{ext}"
        # Close all running Excel application before save
        pythoncom.CoInitialize()
        try:
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
                excel.Quit()
            except Exception:
                pass
        finally:
            pythoncom.CoUninitialize()
        time.sleep(1)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Session"
        ws.append(["File Path", "Sheet Name", "Cell Address"])
        for _, path, sheet, cell in selected: ws.append([path, sheet, cell])
        wb.save(file_path_with_ts)
        messagebox.showinfo("Success", f"Session saved at:\n{file_path_with_ts}")
        # Reload session (open saved files)
        self.load_session_from_path(file_path_with_ts)


    def load_session(self):
        current_files, _, _, _ = self.get_open_excel_files()
        if current_files:
            messagebox.showwarning("Warning", "Please close all currently open Excel files before loading a session.")
            self.show_names()
            return
        file_path = filedialog.askopenfilename(
            title="Load Session", filetypes=[("Excel Session", "*.xlsx"), ("All Files", "*.*")]
        )
        if not file_path or not os.path.exists(file_path): return
        def thread_job():
            pythoncom.CoInitialize()
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                valid_rows = [r for r in rows if r and r[0] and os.path.exists(r[0])]
                if not valid_rows:
                    self.root.after(0, lambda: messagebox.showwarning("Warning", "No valid file paths found."))
                    return
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True
                excel.AskToUpdateLinks = False
                for r in valid_rows:
                    path, sheet, cell = (r[0], r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)
                    try:
                        wb_xl = excel.Workbooks.Open(Filename=path, UpdateLinks=0)
                        if sheet:
                            try:
                                sht = wb_xl.Sheets(sheet)
                                sht.Activate()
                                if cell: sht.Range(cell).Select()
                            except Exception: pass
                    except Exception: pass
                excel.AskToUpdateLinks = True
                self.root.after(0, lambda: messagebox.showinfo("Complete", f"{len(valid_rows)} file(s) opened."))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Error loading session:\n{str(e)}"))
            finally:
                pythoncom.CoUninitialize()
                self.root.after(200, self.show_names)
        threading.Thread(target=thread_job, daemon=True).start()

    def load_session_from_path(self, file_path):
        def thread_job():
            pythoncom.CoInitialize()
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                valid_rows = [r for r in rows if r and r[0] and os.path.exists(r[0])]
                if not valid_rows:
                    self.root.after(0, lambda: messagebox.showwarning("Warning", "No valid file paths found."))
                    return
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True
                excel.AskToUpdateLinks = False
                for r in valid_rows:
                    path, sheet, cell = (r[0], r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)
                    try:
                        wb_xl = excel.Workbooks.Open(Filename=path, UpdateLinks=0)
                        if sheet:
                            try:
                                sht = wb_xl.Sheets(sheet)
                                sht.Activate()
                                if cell: sht.Range(cell).Select()
                            except Exception: pass
                    except Exception: pass
                excel.AskToUpdateLinks = True
                self.root.after(0, lambda: messagebox.showinfo("Complete", f"{len(valid_rows)} file(s) opened."))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Error loading session:\n{str(e)}"))
            finally:
                pythoncom.CoUninitialize()
                self.root.after(200, self.show_names)
        threading.Thread(target=thread_job, daemon=True).start()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelSessionManagerApp(root)
    root.mainloop()
